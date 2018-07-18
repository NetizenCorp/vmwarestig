from openpyxl import Workbook



class stig_report(object):

    def __init__(self):
        self.wb = Workbook()
        # grab the active worksheet
        self.ws = self.wb.active
        self.row_count = 1

    def write_row(self, vuln_num, vuln_output):
        self.ws['A' + str(self.row_count)] = vuln_num
        self.ws['B' + str(self.row_count)] = vuln_output
        self.row_count += 1

    def close_report(self):
        self.wb.save("sample.xlsx")



s = stig_report()
s.write_row('V-4444', 'test this')
s.close_report()

# Data can be assigned directly to cells
#ws['A1'] = 42

# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()

# Save the file
#wb.save("sample.xlsx")
